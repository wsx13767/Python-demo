#! python3
# blankRowInserter.py - insert blank row
import openpyxl, sys
from openpyxl.utils import get_column_letter
if len(sys.argv) == 4 :
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.active
    maxrow = sheet.max_row
    maxcol = sheet.max_column
    cellArray = {}
    for i in range(int(sys.argv[1]), maxrow + 1) :
        for j in range(1, maxcol+1) :
            cellArray[get_column_letter(j) + str(i + int(sys.argv[2]))] = sheet[get_column_letter(j) + str(i)].value
            sheet[get_column_letter(j) + str(i)] = ''
    for key in cellArray :
        sheet[key] = cellArray[key]
    wb.save(sys.argv[3])
