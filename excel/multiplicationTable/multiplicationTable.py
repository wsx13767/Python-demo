#! python3
# multiplicationTable.py - NXN
import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string
wb = openpyxl.Workbook()
sheet = wb.active
sheet.freeze_panes = 'B2'
fontObj = Font(bold=True)
if len(sys.argv) == 2 :
    num = int(sys.argv[1])
    for i in range(1, num + 1) :
        sheet.cell(row=i+1,column=1).value=i
        sheet.cell(row=i+1,column=1).font = fontObj
        sheet.cell(row=1,column=i+1).value=i
        sheet.cell(row=1,column=i+1).font = fontObj
        for j in range(2, num + 2) :
		# ex sheet[B2] = '=B1*A2'
            sheet[get_column_letter(i+1) + str(j)] = '=' + get_column_letter(i+1)+str(1)+'*A'+str(j)
        
            
wb.save('multiplicationTable.xlsx')
